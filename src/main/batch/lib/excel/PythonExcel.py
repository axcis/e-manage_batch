# coding: UTF-8
'''
PythonExcel
Excel操作ライブラリ

@author: takanori_gozu
'''
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Side, Border
from openpyxl.styles.alignment import Alignment

class PythonExcel:

    excel = None
    sheet = None

    '''
    コンストラクタ
    '''
    def __init__(self):
        self.excel = openpyxl.Workbook()
        self.sheet = self.excel.active

    '''
    シートのリネーム
    '''
    def rename(self, sheetName):
        self.sheet.title = sheetName

    '''
    シートの追加(先頭)
    '''
    def addSheetFirst(self, sheetName):
        self.excel.create_sheet(sheetName, 0)

    '''
    シートの追加(末尾)
    '''
    def addSheetLast(self, sheetName):
        self.excel.create_sheet(sheetName)

    '''
    列幅の設定
    '''
    def setCellWidth(self, col, width):
        self.sheet.column_dimensions[col].width = width

    '''
    行高の設定
    '''
    def setRowHeight(self, no, height):
        self.sheet.row_dimensions[no].height = height

    '''
    ページ設定(A3)
    '''
    def setPageA3(self):
        self.sheet.page_setup.paperSize = self.sheet.PAPERSIZE_A3

    '''
    ページ設定(A4)
    '''
    def setPageA4(self):
        self.sheet.page_setup.paperSize = self.sheet.PAPERSIZE_A4

    '''
    1ページに収める
    '''
    def setFitToPage(self):
        self.sheet.page_setup.fitToPage = True

    '''
    フォントの変更
    '''
    def changeFont(self, range, fontName):
        font = Font(name=fontName)
        self.sheet[range].font = font

    '''
    フォントサイズの変更
    '''
    def changeSize(self, range, size):
        font = Font(size=size)
        self.sheet[range].font = font

    '''
    セルの背景色
    '''
    def changeCellBackColor(self, range, color, type = 'solid'):
        fill = PatternFill(patternType=type, start_color=color, end_color=color)
        self.sheet[range].fill = fill

    '''
    セルの背景色(複数セル)
    '''
    def changeCellBackColorMulti(self, range, color, type = 'solid'):
        startRange = range.split(':')[0]
        endRange = range.split(':')[1]
        fill = PatternFill(patternType=type, start_color=color, end_color=color)
        for rows in self.sheet[startRange : endRange]:
            for cell in rows:
                self.sheet[cell.coordinate].fill = fill

    '''
    セルの文字色
    '''
    def changeCellFrontColor(self, range, color):
        font = Font(color=color)
        self.sheet[range].font = font

    '''
    セルの文字色(複数セル)
    '''
    def changeCellFrontColorMulti(self, range, color):
        startRange = range.split(':')[0]
        endRange = range.split(':')[1]
        font = Font(color=color)
        for rows in self.sheet[startRange : endRange]:
            for cell in rows:
                self.sheet[cell.coordinate].font = font

    '''
    セルの書式設定(縦位置、横位置、セル内改行)
    '''
    def setAlignment(self, range, vertical = 'top', horizon = 'left', wrapText = False):
        self.sheet[range].alignment = Alignment(vertical=vertical, horizontal=horizon, wrapText=wrapText)

    '''
    セルの書式設定(縦位置、横位置、セル内改行)複数セル
    '''
    def setAlignmentMulti(self, range, vertical = 'top', horizon = 'left', wrapText = False):
        startRange = range.split(':')[0]
        endRange = range.split(':')[1]
        for rows in self.sheet[startRange : endRange]:
            for cell in rows:
                self.sheet[cell.coordinate].alignment = Alignment(vertical=vertical, horizontal=horizon, wrapText=wrapText)

    '''
    セルの結合
    '''
    def mergeCell(self, range):
        self.sheet.merge_cells(range)

    '''
    セルの罫線
    '''
    def setBorder(self, range, style = 'thin', color = '000000', place = 'all'):
        side = Side(style=style, color=color)
        if place == 'all':
            border = Border(top=side, bottom=side, left=side, right=side)
        elif place == 'top':
            border = Border(top=side)
        elif place == 'bottom':
            border = Border(bottom=side)
        elif place == 'left':
            border = Border(left=side)
        elif place == 'right':
            border = Border(right=side)

        self.sheet[range].border = border

    '''
    セルの罫線(複数セル)
    '''
    def setBorderMulti(self, range, style = 'thin', color = '000000', place = 'all'):
        side = Side(style=style, color=color)
        if place == 'all':
            border = Border(top=side, bottom=side, left=side, right=side)
        elif place == 'top':
            border = Border(top=side)
        elif place == 'bottom':
            border = Border(bottom=side)
        elif place == 'left':
            border = Border(left=side)
        elif place == 'right':
            border = Border(right=side)

        startRange = range.split(':')[0]
        endRange = range.split(':')[1]
        for rows in self.sheet[startRange : endRange]:
            for cell in rows:
                self.sheet[cell.coordinate].border = border

    '''
    セルに値を代入する(A1方式)
    '''
    def setValueA1(self, range, value):
        self.sheet[range] = value

    '''
    セルに値を代入する(R1C1方式)
    '''
    def setValueR1C1(self, col, row, value):
        self.sheet.cell(row=row, column=col).value = value

    '''
    印刷範囲の設定
    '''
    def setPrintArea(self, range):
        self.sheet.print_area = range

    '''
    余白設定
    '''
    def setMargin(self, top = 0, bottom = 0, left = 0, right = 0, header = 0, footer = 0):
        self.sheet.page_margins.top = top
        self.sheet.page_margins.bottom = bottom
        self.sheet.page_margins.left = left
        self.sheet.page_margins.right = right
        self.sheet.page_margins.header = header
        self.sheet.page_margins.footer = footer

    '''
    保存
    '''
    def save(self, filePath):
        self.excel.save(filePath)